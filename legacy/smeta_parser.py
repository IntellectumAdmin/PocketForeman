# -*- coding: utf-8 -*-
"""
Универсальный парсер смет
Поддерживает: PDF, Excel (XLSX, XLS)
"""

import os
import re
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from openpyxl import load_workbook
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def parse_smeta(file_path: str) -> List[Dict[str, Any]]:
    """
    Автоматически определяет формат и парсит смету
    
    Args:
        file_path: путь к файлу сметы
        
    Returns:
        Список работ
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    ext = Path(file_path).suffix.lower()
    
    if ext == '.pdf':
        return parse_pdf_smeta(file_path)
    elif ext in ['.xlsx', '.xls']:
        return parse_excel_smeta(file_path)
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}")


def parse_pdf_smeta(pdf_path: str) -> List[Dict[str, Any]]:
    """
    Парсит PDF-смету (формат КС - Казахстан)
    """
    if not HAS_PDF:
        raise ImportError("Установи pdfplumber: pip install pdfplumber")
    
    works = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            
            for table in tables:
                if not table:
                    continue
                
                # Ищем строки с кодами работ (формат: XXXX-XXXX-XXXX)
                for row in table:
                    if not row or len(row) < 3:
                        continue
                    
                    # Ищем код работы
                    code = None
                    for cell in row[:3]:
                        if cell and re.match(r'\d{4}-\d{4}-\d{4}', str(cell)):
                            code = str(cell).strip()
                            break
                    
                    if not code:
                        continue
                    
                    # Извлекаем данные
                    work = {
                        "code": code,
                        "name": "",
                        "unit": "",
                        "volume": 0.0,
                        "price": 0.0,
                    }
                    
                    # Ищем название работы (обычно в следующих ячейках)
                    for i, cell in enumerate(row):
                        if not cell:
                            continue
                        
                        cell_str = str(cell).strip()
                        
                        # Название (длинный текст)
                        if len(cell_str) > 15 and not work["name"]:
                            work["name"] = cell_str
                        
                        # Единица измерения
                        if cell_str in ["м³", "м²", "м", "шт", "т", "п.м.", "тыс.шт"]:
                            work["unit"] = cell_str
                        
                        # Объём (число)
                        if re.match(r'^\d+[\s\d]*$', cell_str.replace(' ', '')):
                            try:
                                work["volume"] = float(cell_str.replace(' ', ''))
                            except:
                                pass
                    
                    if work["name"]:
                        works.append(work)
    
    print(f"📄 PDF парсинг: найдено {len(works)} работ")
    return works


def parse_excel_smeta(excel_path: str) -> List[Dict[str, Any]]:
    """
    Парсит Excel-смету
    """
    if not HAS_EXCEL:
        raise ImportError("Установи openpyxl: pip install openpyxl")
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    works = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        
        # Ищем код работы
        code = None
        for cell in row[:3]:
            if cell and re.match(r'\d{4}-\d{4}-\d{4}', str(cell)):
                code = str(cell).strip()
                break
        
        if not code:
            continue
        
        work = {
            "code": code,
            "name": "",
            "unit": "",
            "volume": 0.0,
            "price": 0.0,
        }
        
        for cell in row:
            if not cell:
                continue
            
            cell_str = str(cell).strip()
            
            # Название
            if len(cell_str) > 15 and not work["name"]:
                work["name"] = cell_str
            
            # Единица измерения
            if cell_str in ["м³", "м²", "м", "шт", "т", "п.м."]:
                work["unit"] = cell_str
            
            # Объём
            if isinstance(cell, (int, float)):
                if cell > 0 and work["volume"] == 0:
                    work["volume"] = float(cell)
        
        if work["name"]:
            works.append(work)
    
    print(f"📊 Excel парсинг: найдено {len(works)} работ")
    return works


if __name__ == "__main__":
    # Тест
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        works = parse_smeta(file_path)
        
        print(f"\n{'='*60}")
        print(f"Найдено работ: {len(works)}")
        print(f"{'='*60}\n")
        
        for i, w in enumerate(works[:10], 1):
            print(f"{i}. {w['code']}: {w['name']}")
            print(f"   Объём: {w['volume']} {w['unit']}")
            print()
    else:
        print("Использование: python smeta_parser.py <путь_к_файлу>")